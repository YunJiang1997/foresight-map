"""
Run this script whenever the Excel file is updated.
It regenerates public/foresight_data.json from the source spreadsheet.
No visual changes — data only.
"""

import openpyxl
import json
import os

EXCEL_PATH = (
    r"C:\Users\jiang\OneDrive - Pernod Ricard"
    r"\Global Tech - Technology Innovation - Documents"
    r"\1. Foresight Studies"
    r"\1.5 Technologies Mapping and Chains-of-Value"
    r"\Technologies Mapping Resources"
    r"\PR_Technology_Foresight_Mapping_Bold_Horizons_TO REVIEW.xlsx"
)

SHEET_NAME  = "Foresight Mapping (Updated)"
EXCLUDE_ROWS = range(98, 109)   # rows 98–108 inclusive

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "public", "foresight_data.json")

def s(v):
    return str(v).strip() if v and str(v).strip() not in ('None', 'N/A', 'null', 'N/A ') else ""

def build_tree(ws):
    tree = {}
    for row in range(2, ws.max_row + 1):
        if row in EXCLUDE_ROWS:
            continue
        # Columns: A=1, B=2, ..., M=13
        domain    = ws.cell(row=row, column=1).value
        portfolio = ws.cell(row=row, column=2).value
        pf        = ws.cell(row=row, column=3).value
        pf_desc   = ws.cell(row=row, column=4).value
        tech      = ws.cell(row=row, column=5).value
        tech_desc = ws.cell(row=row, column=6).value
        source    = ws.cell(row=row, column=7).value
        comment   = ws.cell(row=row, column=8).value   # H — Comment (Implications)
        horizon   = ws.cell(row=row, column=9).value
        stage     = ws.cell(row=row, column=10).value
        maturity  = ws.cell(row=row, column=11).value  # K — Maturity at Pernod Ricard
        cases     = ws.cell(row=row, column=12).value  # L — Successful Cases in Other Companies
        science   = ws.cell(row=row, column=13).value  # M — Scientific Support

        if not domain:
            continue

        tree.setdefault(domain, {"name": domain, "portfolios": {}})
        tree[domain]["portfolios"].setdefault(portfolio, {"name": portfolio, "productFamilies": {}})
        tree[domain]["portfolios"][portfolio]["productFamilies"].setdefault(pf, {
            "name": pf,
            "description": s(pf_desc),
            "technologies": []
        })
        tree[domain]["portfolios"][portfolio]["productFamilies"][pf]["technologies"].append({
            "name":                        s(tech),
            "description":                 s(tech_desc),
            "source":                      s(source),
            "implications":                s(comment),
            "horizon":                     s(horizon),
            "stage":                       s(stage),
            "maturityAtPR":                s(maturity),
            "successfulCasesOtherCompanies": s(cases),
            "scientificSupport":           s(science),
        })
    return tree

def tree_to_list(tree):
    result = []
    for domain_val in tree.values():
        portfolios = []
        for port_val in domain_val["portfolios"].values():
            product_families = list(port_val["productFamilies"].values())
            portfolios.append({"name": port_val["name"], "productFamilies": product_families})
        result.append({"name": domain_val["name"], "portfolios": portfolios})
    return result

if __name__ == "__main__":
    print(f"Reading: {EXCEL_PATH}")
    wb   = openpyxl.load_workbook(EXCEL_PATH)
    ws   = wb[SHEET_NAME]

    tree = build_tree(ws)
    data = tree_to_list(tree)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    domains    = len(data)
    portfolios = sum(len(d["portfolios"]) for d in data)
    pfs        = sum(len(p["productFamilies"]) for d in data for p in d["portfolios"])
    techs      = sum(len(pf["technologies"]) for d in data for p in d["portfolios"] for pf in p["productFamilies"])

    print(f"Done -> {OUTPUT_PATH}")
    print(f"  {domains} Domains · {portfolios} Portfolios · {pfs} Product Families · {techs} Technologies")
    print(f"  (rows 98–108 excluded)")
